# _*_ coding: utf-8 _*_
__author__ = 'pblh123'
#Find special file with python3

import os,sys
import time
import csv
import xlsxwriter
import openpyxl

#Get create time
def Ctime(File):
    ctime=time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(os.path.getctime(File)))
    return ctime
#Last modified time
def Mtime(File):
    mtime=time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(os.path.getmtime(File)))
    return mtime

#Get File size
def GetSize(File):
    Size=os.stat(File).st_size
    return Size

#Get Last access time
def Atime(File):
    atime=time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(os.path.getatime(File)))
    return atime

# create excel file
def ToExcel(Path):
    OutputEcevelFile=Path+'\\'+"Collecter_files.xlsx"
    FileTypes=[]
    for root,dirs,files in os.walk(Path):
        for file in files:
            if os.path.splitext(file)[1][1:] in FileTypes:

                continue
            else:
                FileTypes.append(os.path.splitext(file)[1][1:])
    print(FileTypes)
    # create excel file
    workbook=xlsxwriter.Workbook(OutputEcevelFile)
    bold=workbook.add_format({'bold':1})
    for i in FileTypes:
        if i != '':
            try:
                worksheet=workbook.add_worksheet(i)
                row=0
                worksheet.write(row,0,"文件名",bold)
                worksheet.write(row,1,"文件大小",bold)
                worksheet.write(row,2,"文件路径",bold)
                worksheet.write(row,3,"文件创建时间",bold)
                worksheet.write(row,4,"文件修改时间",bold)
                worksheet.write(row,5,"最后访问时间",bold)
                row+=1
            except:
                print(i+'is unsupport format!')
                continue
            for root,dirs,files in os.walk(Path):
                for file in files:
                    if os.path.splitext(file)[1][1:] ==i:
                        try:
                            File=os.path.join(root,file)
                            ctime=Ctime(File)
                            mtime=Mtime(File)
                            atime=Atime(File)
                            fsize=GetSize(File)
                            worksheet.write(row,0,file)
                            worksheet.write(row,1,fsize)
                            worksheet.write(row,2,root)
                            worksheet.write(row,3,ctime)
                            worksheet.write(row,4,mtime)
                            worksheet.write(row,5,atime)
                            Str= file+'\t'+str(fsize)+'\t'+root+'\t'+ctime+'\t'+mtime+'\t'+atime+'\n'
                            print(Str)
                            row+=1
                        except:
                            print('Here is wrong!')
    workbook.close()
    ExcelFile=openpyxl.load_workbook(OutputEcevelFile)
    ContentTable=ExcelFile.create_sheet('content',0)
    row0=1
    No0=1
    ContentTable['A1']='序号'
    ContentTable['B1']='文件类型'
    ContentTable['C1']='文件数量'
    for j in FileTypes:
        if j != '':
            try:
                row0+=1
                links='#'+j+'!A1'
                ExcelTables=ExcelFile.get_sheet_by_name(j)
                TableRowsLen=ExcelTables.max_row
                ContentTable.cell(row=row0,column=1).value = No0
                ContentTable.cell(row=row0,column=2).value = '=HYPERLINK("{}","{}")'.format(links,j)
                ContentTable.cell(row=row0,column=3).value = TableRowsLen -1
                No0+=1
            except:
                print('Here is wrong!')
    ExcelFile.save(OutputEcevelFile)

if __name__ == '__main__':
    print('Please input path:')
    Path=input()
    ToExcel(Path)
